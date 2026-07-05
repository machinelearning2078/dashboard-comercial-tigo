# ============================================================
# DASHBOARD COMERCIAL AVANZADO - TIGO
# Con RFM, heatmaps, pronósticos y Excel profesional
# ============================================================

import pandas as pd
import numpy as np
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from statsmodels.tsa.arima.model import ARIMA
from statsmodels.tsa.stattools import adfuller
from sklearn.metrics import mean_absolute_error, mean_squared_error
import warnings
warnings.filterwarnings('ignore')

# Configuración de página
st.set_page_config(page_title="Dashboard Comercial - Tigo", layout="wide", initial_sidebar_state="expanded")
st.title("📊 Análisis Comercial Interactivo")
st.markdown("---")

# ============================================================
# 1. FUNCIONES DE CARGA CON FILTRO POR AÑO
# ============================================================
@st.cache_data
def get_available_years():
    """Obtiene los años disponibles en los datos (rápido, solo lee fechas)"""
    df = pd.read_parquet("facturacion.parquet")
    df['F. Emisión'] = pd.to_datetime(df['F. Emisión'], dayfirst=True, errors='coerce')
    years = sorted(df['F. Emisión'].dt.year.dropna().unique(), reverse=True)
    return years

@st.cache_data
def load_data(selected_year=2025):
    """Carga datos filtrados por año para ahorrar memoria"""
    ruta = "facturacion.parquet"
    df = pd.read_parquet(ruta)
    
    # Limpiar nombres
    df.columns = df.columns.str.strip()
    
    # Convertir fechas (formato DD/MM/YYYY)
    date_cols = ['F. Emisión', 'F. Venc.', 'F. Liquidación']
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], dayfirst=True, errors='coerce')
    
    # 🔥 FILTRO POR AÑO: carga solo el año seleccionado (por defecto 2025)
    if selected_year != 0:  # 0 = "Todos"
        df = df[df['F. Emisión'].dt.year == selected_year]
    
    # Convertir numéricos
    num_cols = ['Precio Unitario', 'Precio Por Renglon', 'Cant. Cajas']
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    # Crear columnas de tiempo
    if 'F. Emisión' in df.columns:
        df['Mes_Año'] = df['F. Emisión'].dt.to_period('M').astype(str)
        df['Año'] = df['F. Emisión'].dt.year
        df['Mes'] = df['F. Emisión'].dt.month
        df['Semana'] = df['F. Emisión'].dt.isocalendar().week
        df['Día'] = df['F. Emisión'].dt.date
    
    # Eliminar nulos críticos
    df = df.dropna(subset=['Precio Por Renglon', 'Cant. Cajas', 'F. Emisión'])
    
    # Calcular Dropsize: total de cajas por documento (N° Documento)
    dropsize_per_doc = df.groupby('N° Documento')['Cant. Cajas'].sum().reset_index()
    dropsize_per_doc.columns = ['N° Documento', 'Dropsize']
    df = df.merge(dropsize_per_doc, on='N° Documento', how='left')
    
    return df

# ============================================================
# 2. FILTROS DINÁMICOS (con selector de año en sidebar)
# ============================================================
st.sidebar.header("🔍 Filtros")

# Obtener años disponibles
available_years = get_available_years()
year_options = [0] + available_years  # 0 = "Todos"
year_labels = ["Todos"] + [str(y) for y in available_years]

selected_year_label = st.sidebar.selectbox("📅 Año", year_labels, index=1 if 2025 in available_years else 0)
selected_year = 0 if selected_year_label == "Todos" else int(selected_year_label)

# Cargar datos con el año seleccionado
df = load_data(selected_year)
st.sidebar.success(f"✅ Datos cargados: {len(df):,} registros")

# Filtros de fechas (dentro del año seleccionado)
if df.empty:
    st.warning("No hay datos para el año seleccionado.")
    st.stop()

min_date = df['F. Emisión'].min().date()
max_date = df['F. Emisión'].max().date()

filtro_tiempo = st.sidebar.radio("Agrupación temporal", ["Mensual", "Semanal"], index=0)

col1, col2 = st.sidebar.columns(2)
with col1:
    fecha_inicio = st.date_input("Fecha inicio", min_date, min_value=min_date, max_value=max_date)
with col2:
    fecha_fin = st.date_input("Fecha fin", max_date, min_value=min_date, max_value=max_date)

# Filtros con buscador
clientes_opciones = df['Nombre Cliente'].dropna().unique()
clientes = st.sidebar.multiselect("Cliente (Nombre)", options=clientes_opciones, default=[], help="Escribe para buscar")

categorias_opciones = df['Categoría'].dropna().unique()
categorias = st.sidebar.multiselect("Categoría", options=categorias_opciones, default=[], help="Escribe para buscar")

productos_opciones = df['Descripción'].dropna().unique()
productos = st.sidebar.multiselect("Producto (Descripción)", options=productos_opciones, default=[], help="Escribe para buscar")

lineas_opciones = df['Línea'].dropna().unique()
lineas_opciones = [l for l in lineas_opciones if l not in ['Repacking', 'Zservicios']]
lineas = st.sidebar.multiselect("Línea", options=lineas_opciones, default=[], help="Escribe para buscar")

# Aplicar filtros
mask = (df['F. Emisión'].dt.date >= fecha_inicio) & (df['F. Emisión'].dt.date <= fecha_fin)
if clientes:
    mask &= df['Nombre Cliente'].isin(clientes)
if categorias:
    mask &= df['Categoría'].isin(categorias)
if productos:
    mask &= df['Descripción'].isin(productos)
if lineas:
    mask &= df['Línea'].isin(lineas)

df_filtrado = df.loc[mask]

if df_filtrado.empty:
    st.warning("No hay datos con los filtros seleccionados. Ajusta los criterios.")
    st.stop()

# Excluir cliente "CENTRAL GANADERO LA TORA" (productos defectuosos)
df_filtrado = df_filtrado[~df_filtrado['Nombre Cliente'].str.contains("CENTRAL GANADERO LA TORA", case=False, na=False)]

st.sidebar.write(f"**Registros:** {len(df_filtrado):,}")
st.sidebar.write(f"**Periodo:** {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}")

# ============================================================
# 3. MÉTRICAS GLOBALES
# ============================================================
total_ventas = df_filtrado['Precio Por Renglon'].sum()
total_cajas = df_filtrado['Cant. Cajas'].sum()
num_clientes = df_filtrado['Código Cliente'].nunique()
num_facturas = df_filtrado['N° Documento'].nunique()
ticket_prom = total_ventas / num_facturas if num_facturas > 0 else 0

def fmt_currency(val):
    return f"${val:,.2f}"

col1, col2, col3, col4 = st.columns(4)
col1.metric("💰 Ventas Netas", fmt_currency(total_ventas))
col2.metric("📦 Cajas Vendidas", f"{total_cajas:,.0f}")
col3.metric("🧾 Ticket Promedio", fmt_currency(ticket_prom))
col4.metric("👥 Clientes Únicos", f"{num_clientes:,}")

st.markdown("---")

# ============================================================
# 4. ORGANIZACIÓN EN PESTAÑAS
# ============================================================
tab1, tab2, tab3, tab4 = st.tabs([
    "📈 Evolución", 
    "🗺️ Heatmaps", 
    "🔮 Pronóstico", 
    "👤 Clientes RFM"
])

# ============================================================
# FUNCIÓN PARA GENERAR TÍTULOS DINÁMICOS
# ============================================================
def generar_titulo(base, filtros_aplicados):
    if filtros_aplicados:
        return f"{base} - {', '.join(filtros_aplicados[:3])}{' ...' if len(filtros_aplicados)>3 else ''}"
    else:
        return base

# ============================================================
# TAB 1: EVOLUCIÓN TEMPORAL
# ============================================================
with tab1:
    st.subheader("Evolución de Ventas y Cajas")
    if filtro_tiempo == "Mensual":
        df_agg = df_filtrado.groupby('Mes_Año').agg({
            'Precio Por Renglon': 'sum',
            'Cant. Cajas': 'sum'
        }).reset_index()
        df_agg['Fecha'] = pd.to_datetime(df_agg['Mes_Año'] + '-01')
        df_agg = df_agg.sort_values('Fecha')
        df_agg['Fecha_str'] = df_agg['Fecha'].dt.strftime('%b %Y')
    else:
        df_filtrado['Semana_Año'] = df_filtrado['F. Emisión'].dt.to_period('W').astype(str)
        df_agg = df_filtrado.groupby('Semana_Año').agg({
            'Precio Por Renglon': 'sum',
            'Cant. Cajas': 'sum'
        }).reset_index()
        df_agg['Fecha'] = pd.to_datetime(df_agg['Semana_Año'] + '-1', format='%Y-W%W-%w', errors='coerce')
        df_agg = df_agg.dropna(subset=['Fecha']).sort_values('Fecha')
        df_agg['Fecha_str'] = df_agg['Fecha'].dt.strftime('Sem %W %b %Y')
    
    filtros_nombres = []
    if clientes: filtros_nombres.append(f"Cliente: {', '.join(clientes[:2])}")
    if productos: filtros_nombres.append(f"Producto: {', '.join(productos[:2])}")
    if categorias: filtros_nombres.append(f"Categoría: {', '.join(categorias[:2])}")
    if lineas: filtros_nombres.append(f"Línea: {', '.join(lineas[:2])}")
    titulo_evol = generar_titulo(f"Ventas y Cajas por {filtro_tiempo}", filtros_nombres)
    
    fig_evol = make_subplots(specs=[[{"secondary_y": True}]])
    fig_evol.add_trace(go.Bar(x=df_agg['Fecha_str'], y=df_agg['Precio Por Renglon'],
                              name="Ventas", marker_color='#d62728'),
                      secondary_y=False)
    fig_evol.add_trace(go.Scatter(x=df_agg['Fecha_str'], y=df_agg['Cant. Cajas'],
                                  name="Cajas", marker_color='#ff7f0e', mode='lines+markers'),
                      secondary_y=True)
    fig_evol.update_layout(title=titulo_evol,
                          xaxis_title="Fecha", height=500,
                          legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
    fig_evol.update_yaxes(title_text="Ventas ($)", secondary_y=False)
    fig_evol.update_yaxes(title_text="Cajas", secondary_y=True)
    fig_evol.update_xaxes(tickangle=-45)
    st.plotly_chart(fig_evol, use_container_width=True)

# ============================================================
# TAB 2: HEATMAPS + GRÁFICOS TOP DINÁMICOS
# ============================================================
with tab2:
    st.subheader("Mapas de Calor por Segmento")
    periodo_texto = f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    filtros_nombres_heat = []
    if clientes: filtros_nombres_heat.append(f"Cliente: {', '.join(clientes[:2])}")
    if productos: filtros_nombres_heat.append(f"Producto: {', '.join(productos[:2])}")
    if categorias: filtros_nombres_heat.append(f"Categoría: {', '.join(categorias[:2])}")
    if lineas: filtros_nombres_heat.append(f"Línea: {', '.join(lineas[:2])}")
    
    def plot_heatmap_full(df, index_col, value_col='Precio Por Renglon', title='', color_scale='Reds'):
        pivot = df.pivot_table(index=index_col, columns='Mes_Año', values=value_col, aggfunc='sum', fill_value=0)
        meses_ordenados = sorted(pivot.columns, key=lambda x: pd.to_datetime(x + '-01'))
        pivot = pivot[meses_ordenados]
        pivot.columns = [pd.to_datetime(c + '-01').strftime('%b %Y') for c in pivot.columns]
        fig = px.imshow(pivot,
                        labels=dict(x="Mes-Año", y=index_col, color="Ventas ($)"),
                        title=f"{title} ({periodo_texto})",
                        color_continuous_scale=color_scale,
                        aspect="auto",
                        text_auto='.1s')
        fig.update_layout(height=450, xaxis_tickangle=-45)
        return fig
    
    if 'Región' in df_filtrado.columns:
        st.plotly_chart(plot_heatmap_full(df_filtrado, 'Región', title='Ventas por Mes y Región'), use_container_width=True)
    if 'Línea' in df_filtrado.columns:
        st.plotly_chart(plot_heatmap_full(df_filtrado, 'Línea', title='Ventas por Mes y Línea'), use_container_width=True)
    if 'Categoría' in df_filtrado.columns:
        st.plotly_chart(plot_heatmap_full(df_filtrado, 'Categoría', title='Ventas por Mes y Categoría'), use_container_width=True)
    
    st.write("**Mapa de Calor: Dropsize Promedio por Mes y Región**")
    if 'Región' in df_filtrado.columns and 'Dropsize' in df_filtrado.columns:
        df_dropsize = df_filtrado.groupby(['Mes_Año', 'Región'], as_index=False)['Dropsize'].mean()
        pivot_dropsize = df_dropsize.pivot_table(index='Región', columns='Mes_Año', values='Dropsize', fill_value=0)
        meses_ordenados = sorted(pivot_dropsize.columns, key=lambda x: pd.to_datetime(x + '-01'))
        pivot_dropsize = pivot_dropsize[meses_ordenados]
        pivot_dropsize.columns = [pd.to_datetime(c + '-01').strftime('%b %Y') for c in pivot_dropsize.columns]
        fig_heat_dropsize = px.imshow(
            pivot_dropsize,
            labels=dict(x="Mes-Año", y="Región", color="Dropsize Prom."),
            title=f"Dropsize Promedio por Mes y Región ({periodo_texto})",
            color_continuous_scale='Reds',
            aspect="auto",
            text_auto='.1s'
        )
        fig_heat_dropsize.update_layout(height=450, xaxis_tickangle=-45)
        st.plotly_chart(fig_heat_dropsize, use_container_width=True)
    else:
        st.info("No hay datos de Región o Dropsize para este heatmap.")
    
    st.write("---")
    st.subheader("📊 Top Dinámico por Ticket Promedio y Dropsize")
    
    top_n = st.slider("Número de elementos en el top", min_value=3, max_value=10, value=5, key="top_n")
    
    top_ticket = df_filtrado.groupby('Nombre Cliente').agg({
        'Precio Por Renglon': 'sum',
        'N° Documento': 'nunique'
    }).reset_index()
    top_ticket['Ticket Promedio'] = top_ticket['Precio Por Renglon'] / top_ticket['N° Documento']
    top_ticket = top_ticket.sort_values('Ticket Promedio', ascending=False).head(top_n)
    
    titulo_ticket = generar_titulo(f"Top {top_n} Clientes por Ticket Promedio", filtros_nombres_heat)
    fig_top_ticket = px.bar(top_ticket, x='Nombre Cliente', y='Ticket Promedio',
                            title=titulo_ticket,
                            text=top_ticket['Ticket Promedio'].apply(lambda x: f"${x:,.1f}"),
                            color='Ticket Promedio', color_continuous_scale='Blues')
    fig_top_ticket.update_traces(textposition='outside')
    fig_top_ticket.update_layout(height=400, xaxis_tickangle=-45)
    st.plotly_chart(fig_top_ticket, use_container_width=True)
    
    with st.expander("📋 Detalle del cálculo del Ticket Promedio"):
        st.markdown(f"""
        **Ticket Promedio** = Ventas Netas / Número de Facturas  
        - **Ventas Netas**: ${total_ventas:,.2f}  
        - **Número de Facturas**: {num_facturas:,}  
        - **Ticket Promedio**: ${ticket_prom:,.2f}
        """)
        st.caption("El ticket promedio se calcula a nivel global para el período filtrado.")
    
    top_dropsize = df_filtrado.groupby('Nombre Cliente')['Dropsize'].mean().reset_index()
    top_dropsize = top_dropsize.sort_values('Dropsize', ascending=False).head(top_n)
    
    titulo_dropsize = generar_titulo(f"Top {top_n} Clientes por Dropsize Promedio", filtros_nombres_heat)
    fig_top_dropsize = px.bar(top_dropsize, x='Nombre Cliente', y='Dropsize',
                              title=titulo_dropsize,
                              text=top_dropsize['Dropsize'].apply(lambda x: f"{x:,.1f}"),
                              color='Dropsize', color_continuous_scale='Oranges')
    fig_top_dropsize.update_traces(textposition='outside')
    fig_top_dropsize.update_layout(height=400, xaxis_tickangle=-45)
    st.plotly_chart(fig_top_dropsize, use_container_width=True)

# ============================================================
# TAB 3: PRONÓSTICOS
# ============================================================
with tab3:
    st.subheader("Pronóstico de Ventas")
    
    if filtro_tiempo == "Mensual":
        series = df_filtrado.groupby('Mes_Año')['Precio Por Renglon'].sum()
        series.index = pd.to_datetime(series.index + '-01')
        freq = 'MS'
        forecast_horizon = 3
    else:
        df_filtrado['Semana'] = df_filtrado['F. Emisión'].dt.to_period('W')
        series = df_filtrado.groupby('Semana')['Precio Por Renglon'].sum()
        series.index = series.index.to_timestamp()
        freq = 'W'
        forecast_horizon = 6
    
    series = series.asfreq(freq)
    series = series.dropna()
    
    if len(series) < 4:
        st.warning("⚠️ No hay suficientes datos históricos para pronosticar.")
    else:
        # ETS
        try:
            if len(series) >= 12:
                model_ets = ExponentialSmoothing(series, trend='add', seasonal='add', seasonal_periods=12)
            else:
                model_ets = ExponentialSmoothing(series, trend='add', seasonal=None)
            fit_ets = model_ets.fit()
            forecast_ets = fit_ets.forecast(forecast_horizon)
        except:
            forecast_ets = None
        
        # Media Móvil
        ma_window = min(3, len(series)-1)
        if ma_window >= 1:
            ma_forecast = series.rolling(window=ma_window).mean().iloc[-1]
            forecast_ma = pd.Series([ma_forecast] * forecast_horizon, 
                                    index=pd.date_range(start=series.index[-1], periods=forecast_horizon+1, freq=freq)[1:])
        else:
            forecast_ma = None
        
        # ARIMA
        try:
            adf_result = adfuller(series.dropna())
            p_value = adf_result[1]
            d = 0 if p_value < 0.05 else 1
            model_arima = ARIMA(series, order=(1, d, 1))
            fit_arima = model_arima.fit()
            forecast_arima = fit_arima.forecast(steps=forecast_horizon)
            forecast_arima.index = pd.date_range(start=series.index[-1], periods=forecast_horizon+1, freq=freq)[1:]
        except:
            forecast_arima = None
        
        # Validación
        errors = {}
        if len(series) >= 6:
            train = series[:-1]
            test = series[-1:]
            try:
                fit_ets_train = ExponentialSmoothing(train, trend='add', 
                                                     seasonal='add' if len(train)>=12 else None,
                                                     seasonal_periods=12 if len(train)>=12 else None).fit()
                pred_ets = fit_ets_train.forecast(1).iloc[0]
                errors['ETS'] = mean_absolute_error(test, [pred_ets])
            except:
                errors['ETS'] = np.inf
            if ma_window >= 1:
                pred_ma = train.rolling(window=ma_window).mean().iloc[-1]
                errors['MA'] = mean_absolute_error(test, [pred_ma])
            else:
                errors['MA'] = np.inf
            try:
                fit_arima_train = ARIMA(train, order=(1, d, 1)).fit()
                pred_arima = fit_arima_train.forecast(1).iloc[0]
                errors['ARIMA'] = mean_absolute_error(test, [pred_arima])
            except:
                errors['ARIMA'] = np.inf
            
            best_method = min(errors, key=errors.get)
            if best_method == 'ETS':
                best_forecast = forecast_ets
            elif best_method == 'MA':
                best_forecast = forecast_ma
            else:
                best_forecast = forecast_arima
        else:
            best_method = None
            best_forecast = None
        
        # Figura
        fig_forecast = go.Figure()
        fig_forecast.add_trace(go.Scatter(x=series.index, y=series, mode='lines+markers', 
                                          name='Histórico', line=dict(color='#1f77b4', width=2)))
        colors = {'ETS': '#d62728', 'MA': '#2ca02c', 'ARIMA': '#9467bd'}
        if forecast_ets is not None:
            fig_forecast.add_trace(go.Scatter(x=forecast_ets.index, y=forecast_ets, 
                                              mode='lines+markers', name='ETS (Suavizamiento Exp.)', 
                                              line=dict(color=colors['ETS'], dash='dot', width=2)))
        if forecast_ma is not None:
            fig_forecast.add_trace(go.Scatter(x=forecast_ma.index, y=forecast_ma, 
                                              mode='lines+markers', name='Media Móvil', 
                                              line=dict(color=colors['MA'], dash='dot', width=2)))
        if forecast_arima is not None:
            fig_forecast.add_trace(go.Scatter(x=forecast_arima.index, y=forecast_arima, 
                                              mode='lines+markers', name='ARIMA', 
                                              line=dict(color=colors['ARIMA'], dash='dot', width=2)))
        if best_forecast is not None:
            fig_forecast.add_trace(go.Scatter(x=best_forecast.index, y=best_forecast, 
                                              mode='lines+markers', name=f'✅ Mejor ({best_method})', 
                                              line=dict(color='#ff7f0e', width=4)))
        
        filtros_nombres_fc = []
        if clientes: filtros_nombres_fc.append(f"Cliente: {', '.join(clientes[:2])}")
        if productos: filtros_nombres_fc.append(f"Producto: {', '.join(productos[:2])}")
        if categorias: filtros_nombres_fc.append(f"Categoría: {', '.join(categorias[:2])}")
        if lineas: filtros_nombres_fc.append(f"Línea: {', '.join(lineas[:2])}")
        titulo_fc = generar_titulo(f"Pronósticos para los próximos {forecast_horizon} periodos ({filtro_tiempo})", filtros_nombres_fc)
        
        fig_forecast.update_layout(title=titulo_fc,
                                   xaxis_title="Fecha", yaxis_title="Ventas ($)", height=550,
                                   legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        fig_forecast.update_xaxes(tickangle=-45)
        st.plotly_chart(fig_forecast, use_container_width=True)
        
        with st.expander("📊 Comparación de modelos de pronóstico (MAE)"):
            if len(series) >= 6:
                df_errors = pd.DataFrame(list(errors.items()), columns=['Modelo', 'MAE'])
                df_errors['MAE'] = df_errors['MAE'].round(2)
                df_errors['Ranking'] = df_errors['MAE'].rank()
                def get_forecast_info(modelo):
                    if modelo == 'ETS' and forecast_ets is not None:
                        idx = forecast_ets.index[0]
                        if isinstance(idx, pd.Timestamp):
                            fecha_str = idx.strftime('%d/%m/%Y')
                        else:
                            fecha_str = str(idx)
                        return f"{fecha_str}: ${forecast_ets.iloc[0]:,.0f}"
                    elif modelo == 'MA' and forecast_ma is not None:
                        idx = forecast_ma.index[0]
                        if isinstance(idx, pd.Timestamp):
                            fecha_str = idx.strftime('%d/%m/%Y')
                        else:
                            fecha_str = str(idx)
                        return f"{fecha_str}: ${forecast_ma.iloc[0]:,.0f}"
                    elif modelo == 'ARIMA' and forecast_arima is not None:
                        idx = forecast_arima.index[0]
                        if isinstance(idx, pd.Timestamp):
                            fecha_str = idx.strftime('%d/%m/%Y')
                        else:
                            fecha_str = str(idx)
                        return f"{fecha_str}: ${forecast_arima.iloc[0]:,.0f}"
                    else:
                        return "No disponible"
                df_errors['Pronóstico (Fecha)'] = df_errors['Modelo'].apply(get_forecast_info)
                df_errors = df_errors.sort_values('MAE')
                st.dataframe(df_errors, use_container_width=True)
                st.markdown("""
                **MAE (Error Absoluto Medio)**: Mide la precisión del pronóstico. 
                Se calcula como el promedio de las diferencias absolutas entre los valores reales y los pronosticados. 
                **Menor MAE = mejor modelo**.
                """)
            else:
                st.info("No hay suficientes datos para calcular errores de validación.")
            st.markdown("**Técnicas utilizadas:**")
            st.markdown("- **ETS (Suavizamiento Exponencial)**: captura tendencia y estacionalidad.")
            st.markdown("- **Media Móvil**: simple, útil para series sin tendencia fuerte.")
            st.markdown("- **ARIMA**: modelo autorregresivo integrado de media móvil, robusto para series con autocorrelación.")

# ============================================================
# TAB 4: CLIENTES RFM
# ============================================================
with tab4:
    st.subheader("Segmentación RFM de Clientes")
    
    fecha_ref = df_filtrado['F. Emisión'].max()
    rfm = df_filtrado.groupby('Código Cliente').agg({
        'F. Emisión': lambda x: (fecha_ref - x.max()).days,
        'N° Documento': 'nunique',
        'Precio Por Renglon': 'sum'
    }).reset_index()
    rfm.columns = ['Código Cliente', 'Recencia', 'Frecuencia', 'Monto']
    ultima_compra = df_filtrado.groupby('Código Cliente')['F. Emisión'].max().reset_index()
    ultima_compra.columns = ['Código Cliente', 'Última Compra']
    rfm = rfm.merge(ultima_compra, on='Código Cliente', how='left')
    rfm = rfm.merge(df_filtrado[['Código Cliente', 'Nombre Cliente']].drop_duplicates(), on='Código Cliente')
    
    percentil_90 = rfm['Monto'].quantile(0.90)
    percentil_70 = rfm['Frecuencia'].quantile(0.70)
    
    condiciones = [
        (rfm['Monto'] >= percentil_90) & (rfm['Recencia'] <= 15),
        (rfm['Frecuencia'] >= percentil_70) & (rfm['Recencia'] <= 60),
        (rfm['Recencia'] >= 60) & (rfm['Recencia'] <= 179),
        (rfm['Recencia'] >= 180)
    ]
    elecciones = ['VIP', 'Leales', 'En Riesgo', 'Dormidos']
    rfm['Segmento'] = np.select(condiciones, elecciones, default='Otros')
    rfm['Segmento'] = np.where(
        (rfm['Última Compra'] >= fecha_ref - pd.Timedelta(days=60)) & (rfm['Segmento'] == 'Otros'),
        'Nuevos',
        rfm['Segmento']
    )
    
    with st.expander("📋 Criterios de Segmentación RFM (Ver / Ocultar)"):
        st.markdown("#### Segmentos y sus criterios")
        criterios = pd.DataFrame({
            'Segmento': ['🟣 VIP', '🟢 Leales', '🟠 En Riesgo', '🔴 Dormidos', '🟡 Nuevos', '⚪ Otros'],
            'Criterio': [
                f"Monto ≥ ${percentil_90:,.2f} y Recencia ≤ 15 días",
                f"Frecuencia ≥ {percentil_70:.0f} y Recencia ≤ 60 días",
                "Recencia entre 60 y 179 días",
                "Recencia ≥ 180 días",
                "Última compra < 60 días",
                "No cumple con los criterios anteriores"
            ]
        })
        st.dataframe(criterios, use_container_width=True, hide_index=True)
        st.caption("Los umbrales se calculan en base a los datos filtrados.")
    
    st.markdown("---")
    
    st.write("**Distribución de segmentos:**")
    seg_counts = rfm['Segmento'].value_counts().reset_index()
    seg_counts.columns = ['Segmento', 'Clientes']
    
    color_map = {
        'VIP': '#8B008B',
        'Leales': '#228B22',
        'En Riesgo': '#FF8C00',
        'Dormidos': '#DC143C',
        'Nuevos': '#FFD700',
        'Otros': '#708090'
    }
    
    col1, col2 = st.columns([2, 1])
    with col1:
        st.dataframe(seg_counts, use_container_width=True)
    with col2:
        fig_pie = px.pie(seg_counts, values='Clientes', names='Segmento', title='% por Segmento',
                         color='Segmento', color_discrete_map=color_map)
        st.plotly_chart(fig_pie, use_container_width=True)
    
    st.write("**Detalle de clientes por segmento (top 10 por monto):**")
    for seg in ['VIP', 'Leales', 'En Riesgo', 'Dormidos', 'Nuevos', 'Otros']:
        with st.expander(f"{'🟣' if seg=='VIP' else '🟢' if seg=='Leales' else '🟠' if seg=='En Riesgo' else '🔴' if seg=='Dormidos' else '🟡' if seg=='Nuevos' else '⚪'} {seg} ({len(rfm[rfm['Segmento']==seg])} clientes)"):
            top = rfm[rfm['Segmento']==seg].sort_values('Monto', ascending=False).head(10)
            top['Monto'] = top['Monto'].apply(lambda x: f"${x:,.2f}")
            top['Última Compra'] = top['Última Compra'].dt.strftime('%d/%m/%Y')
            st.dataframe(top[['Nombre Cliente', 'Recencia', 'Frecuencia', 'Monto', 'Última Compra']], use_container_width=True)
    
    fig_box = px.box(rfm, x='Segmento', y='Recencia', color='Segmento', 
                     title='Distribución de Recencia por Segmento',
                     color_discrete_map=color_map)
    st.plotly_chart(fig_box, use_container_width=True)

# ============================================================
# EXPORTACIÓN A EXCEL
# ============================================================
st.sidebar.markdown("---")
if st.sidebar.button("📥 Exportar reporte a Excel"):
    with st.spinner("Generando archivo Excel con formato profesional..."):
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import io
        
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "KPIs"
        kpis = [("Ventas Netas", total_ventas), ("Cajas", total_cajas), 
                ("Ticket Promedio", ticket_prom), ("Clientes Únicos", num_clientes)]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        money_fmt = '#,##0.00'
        for i, (k, v) in enumerate(kpis, 1):
            ws1.cell(row=i, column=1, value=k).font = Font(bold=True)
            ws1.cell(row=i, column=2, value=v).number_format = money_fmt if isinstance(v, (float, int)) else 'General'
        ws1.column_dimensions['A'].width = 20
        ws1.column_dimensions['B'].width = 15
        
        ws2 = wb.create_sheet("Evolucion")
        for r in dataframe_to_rows(df_agg, index=False, header=True):
            ws2.append(r)
        for col in range(1, len(df_agg.columns)+1):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            ws2.column_dimensions[get_column_letter(col)].width = 15
        
        ws3 = wb.create_sheet("Clientes_RFM")
        for r in dataframe_to_rows(rfm, index=False, header=True):
            ws3.append(r)
        for col in range(1, len(rfm.columns)+1):
            cell = ws3.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            ws3.column_dimensions[get_column_letter(col)].width = 15
        
        try:
            def fig_to_bytes(fig, width=800, height=400):
                return fig.to_image(format="png", width=width, height=height)
            ws_graf = wb.create_sheet("Graficos")
            img_data = fig_to_bytes(fig_evol)
            img = XLImage(io.BytesIO(img_data))
            img.width = 800
            img.height = 400
            ws_graf.add_image(img, 'A1')
            if 'fig_forecast' in locals():
                img_data2 = fig_to_bytes(fig_forecast)
                img2 = XLImage(io.BytesIO(img_data2))
                img2.width = 800
                img2.height = 400
                ws_graf.add_image(img2, 'A30')
        except Exception as e:
            st.warning(f"No se pudieron insertar imágenes: {e}. El Excel contiene solo datos.")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        st.sidebar.download_button(
            label="📥 Descargar Excel",
            data=output,
            file_name="reporte_comercial.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.sidebar.success("✅ Reporte Excel generado.")

st.sidebar.markdown("---")
st.sidebar.caption("Desarrollado con ❤️ para Tigo")
